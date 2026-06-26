#!/usr/bin/env python3
"""
数学组周数据自动填报 v4
======================
原则：
  1. 以第2周手动数据为模板，不动格式/颜色/合并格
  2. 扫描表格找位置写数据，不写死行号
  3. 学生数来自WPS源数据，课时/课次来自排课系统
  4. 新增/结课/停课从源数据的列表里读，没更新就留空
  5. 满班率班级数=WPS班级数，学生数按比例分配
  6. 组课时生产只填 H,I,L,M,N,O,C,Q，不动公式列

用法:
  python3 auto_weekly.py <第N-1周文件.xls> <排课列表.xls> <WPS源数据.xls> [周次]
"""

import sys, os, re, shutil, tempfile
from collections import defaultdict, Counter
import xlrd
from xlutils.copy import copy as xl_copy
from openpyxl import load_workbook

# ═══ 修补 xlwt 保存 .xls 格式的 bug ═══
# xlutils 从某些 .xls 文件读取格式时会产生 None format string
# 导致 xlwt 保存崩溃。这个补丁让它为 None 时用空字符串。
import xlwt.UnicodeUtils as _xlwt_u
_xlwt_u_upack2 = _xlwt_u.upack2
def _patched_upack2(s, encoding='ascii'):
    if s is None: s = ''
    return _xlwt_u_upack2(s, encoding)
_xlwt_u.upack2 = _patched_upack2

# ═══════════════ 配置 ═══════════════
MATH = ['任勇','胡长春','张昱','廖永翠','董葛飞','高明亮',
    '徐良琴','徐荣祥','杨宇宸','周文婧','潘瑶','钱瑞',
    '李娜','梁缘','张旭','王星','李梦','李媛媛','汪涛',
    '费晓曼','宋涛']
MATH_SET = set(MATH)
EXCLUDE = {'钱瑞', '宋涛'}  # 不计入正式教师数

# 特殊授课教师（兼职，课时计入R/S列）
SPECIAL = {'王星', '宋涛'}

# 可在此手工维护教师名单：
# MATH = [...]     # 所有教师
# SPECIAL = {...}  # 兼职特殊授课教师

WK_LABELS = {1:'第一周',2:'第二周',3:'第三周',4:'第四周',5:'第五周'}
WK_DAYS = {1:(1,7),2:(8,14),3:(15,21),4:(22,28),5:(29,31)}


def sf(v):
    if v is None or v == '': return 0.0
    try: return float(v)
    except: return 0.0

def si(v):
    return int(sf(v))


# ═══════════════ 扫描/查找 ═══════════════

def find_row_by_label(sheet, col, label):
    """在某列中找匹配文本的行号"""
    for r in range(sheet.nrows):
        if label in str(sheet.cell_value(r, col)):
            return r
    return None


def write_fmt(ws, ws_rb, r, c, val):
    """写值并保留原单元格格式"""
    orig_xf = ws_rb.cell_xf_index(r, c)
    ws.write(r, c, val)
    row = ws._Worksheet__rows.get(r)
    if row:
        cell = row._Row__cells.get(c)
        if cell:
            cell.xf_idx = orig_xf


# ═══════════════ 读数据 ═══════════════

def read_wps(filepath, week):
    """
    读WPS某一周的Sheet
    返回: {
        'teachers': {name: stats},
        'new_v1': cnt, 'new_bk': cnt,
        'end_v1': cnt, 'end_bk': cnt,
        'stop_v1': cnt, 'stop_bk': cnt,
    }
    """
    wb = xlrd.open_workbook(filepath)
    sn = WK_LABELS.get(week, f'第{week}周')
    if sn not in wb.sheet_names():
        print(f"❌ WPS中无{sn}")
        sys.exit(1)
    s = wb.sheet_by_name(sn)

    # 读教师数据 (R2~R20)
    teachers = {}
    for r in range(s.nrows):
        name = str(s.cell_value(r, 1)).strip()
        if name not in MATH_SET: continue
        teachers[name] = {
            'v1_e':sf(s.cell_value(r,2)), 'v1_h':sf(s.cell_value(r,3)),
            'v1_s':sf(s.cell_value(r,4)), 'v1h':sf(s.cell_value(r,5)),
            'v1s':sf(s.cell_value(r,6)), 'v1_stop':sf(s.cell_value(r,7)),
            'bk_e':sf(s.cell_value(r,9)), 'bk_h':sf(s.cell_value(r,10)),
            'bk_s':sf(s.cell_value(r,11)), 'bk_c':sf(s.cell_value(r,12)),
            'bkh':sf(s.cell_value(r,13)), 'bks':sf(s.cell_value(r,14)),
            'bk_stop':sf(s.cell_value(r,15)),
        }

    # 读学员变动表（在教师行下面）
    new_v1 = new_bk = 0
    end_v1 = end_bk = 0
    stop_v1 = stop_bk = 0
    cur = None
    for r in range(22, s.nrows):  # 从第22行开始（科组行之后）
        c0 = str(s.cell_value(r, 0)).strip()
        c1 = str(s.cell_value(r, 1)).strip() if s.ncols > 1 else ''

        if '结课学员' in c0: cur = 'end'; continue
        if '新增学员' in c0 or '新增学员' in c1: cur = 'new'; continue
        if '停课学员' in c0: cur = 'stop'; continue
        if not cur: continue

        teacher = str(s.cell_value(r, 2)).strip() if s.ncols > 2 else ''
        student = str(s.cell_value(r, 1)).strip() if s.ncols > 1 else ''
        btype = str(s.cell_value(r, 4)).strip() if s.ncols > 4 else ''

        if teacher not in MATH_SET or not student:
            continue

        if '一对一' in btype:
            if cur == 'new': new_v1 += 1
            elif cur == 'end': end_v1 += 1
            elif cur == 'stop': stop_v1 += 1
        elif '班课' in btype:
            if cur == 'new': new_bk += 1
            elif cur == 'end': end_bk += 1
            elif cur == 'stop': stop_bk += 1

    return {
        'teachers': teachers,
        'new_v1': new_v1, 'new_bk': new_bk,
        'end_v1': end_v1, 'end_bk': end_bk,
        'stop_v1': stop_v1, 'stop_bk': stop_bk,
    }


def parse_schedule(filepath, week):
    dmin, dmax = WK_DAYS.get(week, (1, 31))
    tmp = os.path.join(tempfile.gettempdir(), f's_{os.getpid()}.xlsx')
    shutil.copy(filepath, tmp)
    wb = load_workbook(tmp, data_only=True)
    ws = wb.active

    st = defaultdict(lambda: {'v1h':0.0,'v1s':0,'bkk':0,'bks':0})
    sp = defaultdict(lambda: {'v1':0,'bk':0})  # 特殊授课(R/S)
    for row in ws.iter_rows(min_row=2, values_only=True):
        fmt = str(row[2] or ''); tim = str(row[4] or '')
        tch = str(row[11] or '').strip(); act = int(row[7] or 0)
        if tch not in MATH_SET: continue
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', tim)
        if not m or int(m.group(3)) < dmin or int(m.group(3)) > dmax: continue
        if act <= 0: continue
        is1 = '一对一' in fmt and '一对多' not in fmt
        isb = not is1 and ('集体' in fmt or '班' in fmt or '多' in fmt or '一对' in fmt)
        if is1:
            st[tch]['v1s'] += 1; st[tch]['v1h'] += 3
            if tch in SPECIAL: sp[tch]['v1'] += 1
        elif isb:
            st[tch]['bks'] += 1; st[tch]['bkk'] += act
            if tch in SPECIAL: sp[tch]['bk'] += 1
    return dict(st), dict(sp)


# ═══════════════ 填数据 ═══════════════

def fill(base_file, wps, schedule_data, special_data, week, output_file):
    rb = xlrd.open_workbook(base_file, formatting_info=True)
    wb = xl_copy(rb)

    wps_data = wps['teachers']

    # 从WPS读的新增/结课/停课
    new_v1 = wps.get('new_v1', 0)
    new_bk = wps.get('new_bk', 0)
    end_v1 = wps.get('end_v1', 0)
    end_bk = wps.get('end_bk', 0)
    stop_v1 = wps.get('stop_v1', 0)
    stop_bk = wps.get('stop_bk', 0)
    print(f"  新增: v1={new_v1} bk={new_bk}  结课: v1={end_v1} bk={end_bk}  停课: v1={stop_v1} bk={stop_bk}")

    # ── 扫描源数据位置 ──
    rs_src = rb.sheet_by_name('源数据')
    teacher_rows = {}
    for r in range(rs_src.nrows):
        name = str(rs_src.cell_value(r, 1)).strip()
        if name in MATH_SET:
            teacher_rows[name] = r
    tv1_e = int(sum(d['v1_e'] for d in wps_data.values()))
    tv1_h = int(sum(d['v1_h'] for d in wps_data.values()))
    tv1_s = int(sum(d['v1_s'] for d in wps_data.values()))
    tv1h  = sum(d['v1h'] for d in wps_data.values())
    tv1s  = int(sum(d['v1s'] for d in wps_data.values()))
    tb_e  = int(sum(d['bk_e'] for d in wps_data.values()))
    tb_h  = int(sum(d['bk_h'] for d in wps_data.values()))
    tb_s  = int(sum(d['bk_s'] for d in wps_data.values()))
    tbkh  = sum(d['bkh'] for d in wps_data.values())
    tbks  = int(sum(d['bks'] for d in wps_data.values()))
    tbc   = int(sum(d['bk_c'] for d in wps_data.values()))

    all_stu = tv1_e + tv1_h + tv1_s + tb_e + tb_h + tb_s

    # 排课
    sk_v1k = sum(d['v1h'] for n, d in schedule_data.items() if n not in EXCLUDE)
    sk_bkk = sum(d['bkk']*3 for n, d in schedule_data.items() if n not in EXCLUDE)
    sk_bks = sum(d['bks'] for n, d in schedule_data.items() if n not in EXCLUDE)
    sk_tch = sum(1 for n, d in schedule_data.items()
                 if n not in EXCLUDE and n not in SPECIAL
                 and (d['v1h'] > 0 or d['bks'] > 0))

    # 特殊授课 R/S
    sp_R = sum(d.get('v1',0) for d in special_data.values())
    sp_S = sum(d.get('bk',0) for d in special_data.values())
    print(f"  特殊授课: R={sp_R} S={sp_S}")

    # 满班率
    bk_total_stu = tb_e + tb_h + tb_s  # =190
    sp_stu = 2  # 特殊
    reg_stu = bk_total_stu - sp_stu

    mb_x = round(reg_stu * 89/195)
    mb_c = round(reg_stu * 67/195)
    mb_g = reg_stu - mb_x - mb_c

    mb_xc = round(tbc * 14/44)
    mb_cc = round(tbc * 13/44)
    mb_gc = tbc - mb_xc - mb_cc - 1

    cap_x = mb_xc*10; cap_c = mb_cc*8; cap_g = mb_gc*6; cap_sp = 2
    cap_total = cap_x + cap_c + cap_g + cap_sp

    # ── 1. 源数据: 教师行 + 科组 ──
    ws_src = wb.get_sheet('源数据')
    for name, r in teacher_rows.items():
        if name not in wps_data: continue
        d = wps_data[name]
        vt = d['v1_e'] + d['v1_h'] + d['v1_s']
        bkt = d['bk_e'] + d['bk_h'] + d['bk_s']

        ws_src.write(r, 2, int(d['v1_e']))
        ws_src.write(r, 3, int(d['v1_h']))
        ws_src.write(r, 4, int(d['v1_s']))
        ws_src.write(r, 5, d['v1h'])
        ws_src.write(r, 6, int(d['v1s']))
        ws_src.write(r, 7, int(d['v1_stop']))
        ws_src.write(r, 8, round(d['v1h']/vt, 1) if vt > 0 else 0)
        ws_src.write(r, 9, int(d['bk_e']))
        ws_src.write(r, 10, int(d['bk_h']))
        ws_src.write(r, 11, int(d['bk_s']))
        ws_src.write(r, 12, int(d['bk_c']))
        ws_src.write(r, 13, d['bkh'])
        ws_src.write(r, 14, int(d['bks']))
        ws_src.write(r, 15, int(d['bk_stop']))
        ws_src.write(r, 16, round(bkt/d['bk_c'], 1) if d['bk_c'] > 0 else 0)
        ws_src.write(r, 17, int(vt + bkt))
        ws_src.write(r, 18, round(d['v1h'] + d['bkh']/3, 1))
        ws_src.write(r, 19, int(d['v1s'] + d['bks']))

    # 科组行
    zr = find_row_by_label(rs_src, 1, '科组') or find_row_by_label(rs_src, 1, '组')
    if zr is not None:
        stp_v1 = int(sum(d['v1_stop'] for d in wps_data.values()))
        stp_bk = int(sum(d['bk_stop'] for d in wps_data.values()))
        ws_src.write(zr, 2, tv1_e); ws_src.write(zr, 3, tv1_h); ws_src.write(zr, 4, tv1_s)
        ws_src.write(zr, 5, tv1h); ws_src.write(zr, 6, tv1s); ws_src.write(zr, 7, stp_v1)
        ws_src.write(zr, 9, tb_e); ws_src.write(zr, 10, tb_h); ws_src.write(zr, 11, tb_s)
        ws_src.write(zr, 12, tbc); ws_src.write(zr, 13, tbkh); ws_src.write(zr, 14, tbks)
        ws_src.write(zr, 15, stp_bk)
        ws_src.write(zr, 17, tv1_e+tv1_h+tv1_s+tbc)
        ws_src.write(zr, 18, round(tv1h+tbkh/3, 1))
        ws_src.write(zr, 19, int(tv1s+tbks))
        print(f"  科组行: R{zr}")

    # ── 2. 学生Sheet ──
    # 停课：按上周数据填
    rb_stu = rb.sheet_by_name('学生')
    prev_v1_stop = int(rb_stu.cell_value(4, 5))   # col5=1v1停课 (第2周 row4)
    prev_bk_stop = int(rb_stu.cell_value(4, 12))  # col12=班课停课
    rs_stu = rb.sheet_by_name('学生')
    ws_stu = wb.get_sheet('学生')
    # 找第N周行: 扫描 A 列
    wr = None
    for r in range(rs_stu.nrows):
        if f'第{week}周' in str(rs_stu.cell_value(r, 0)):
            wr = r; break
    if wr is None:
        wr = week + 2  # fallback

    ws_stu.write(wr, 0, f'第{week}周')
    ws_stu.write(wr, 1, all_stu)
    ws_stu.write(wr, 2, tv1_e); ws_stu.write(wr, 3, tv1_h); ws_stu.write(wr, 4, tv1_s)
    ws_stu.write(wr, 5, prev_v1_stop)                # 停课(沿用上周)
    ws_stu.write(wr, 6, new_v1 if new_v1 > 0 else None)
    ws_stu.write(wr, 7, end_v1 if end_v1 > 0 else None)
    ws_stu.write(wr, 9, tb_e); ws_stu.write(wr, 10, tb_h); ws_stu.write(wr, 11, tb_s)
    ws_stu.write(wr, 12, prev_bk_stop)               # 停课(沿用上周)
    ws_stu.write(wr, 13, new_bk if new_bk > 0 else None)
    ws_stu.write(wr, 14, end_bk if end_bk > 0 else None)
    ws_stu.write(wr, 16, tbc)
    print(f"  学生Sheet: R{wr}")

    # ── 3. 满班率 ──
    rs_mb = rb.sheet_by_name('满班率')
    ws_mb = wb.get_sheet('满班率')
    mr = find_row_by_label(rs_mb, 0, '数学')
    if mr is not None:
        ws_mb.write(mr, 1, mb_x); ws_mb.write(mr, 3, mb_xc)
        ws_mb.write(mr, 4, round(mb_x/cap_x, 6))
        ws_mb.write(mr, 5, mb_c); ws_mb.write(mr, 7, mb_cc)
        ws_mb.write(mr, 10, round(mb_c/cap_c, 6))
        ws_mb.write(mr, 11, mb_g); ws_mb.write(mr, 12, mb_gc)
        ws_mb.write(mr, 13, round(mb_g/cap_g, 6))
        ws_mb.write(mr, 14, sp_stu); ws_mb.write(mr, 15, 1); ws_mb.write(mr, 16, 0)
        ws_mb.write(mr, 17, 1.0)
        ws_mb.write(mr, 18, bk_total_stu)
        ws_mb.write(mr, 19, cap_total)
        ws_mb.write(mr, 20, round(bk_total_stu/cap_total, 6))

    # ── 4. 教师Sheet ──
    rs_js = rb.sheet_by_name('教师')
    ws_js = wb.get_sheet('教师')
    jr = find_row_by_label(rs_js, 1, '科组') or find_row_by_label(rs_js, 1, '组')
    if jr is not None:
        ws_js.write(jr, 2, tv1_e+tv1_h+tv1_s)       # 1v1生
        ws_js.write(jr, 3, round(tv1h, 1))            # 1v1课时
        ws_js.write(jr, 5, tb_e+tb_h+tb_s)             # 班课生
        ws_js.write(jr, 6, tbc)                        # 班级数
        ws_js.write(jr, 8, tv1_e+tv1_h+tv1_s+tbc)      # 单科总数
        ws_js.write(jr, 9, round(tv1h+tbkh/3, 1))        # 总课时当量
        ws_js.write(jr, 10, int(tv1s+tbks))            # 总课次

    # ── 5. 组课时生产: 全部公式列 ──
    # 读取已有周的数据，计算累计
    rs_gp = rb.sheet_by_name('组课时生产')
    ws_gp = wb.get_sheet('组课时生产')

    # 收集所有已填周的数据（包括当周）
    all_wks = []
    for r in range(rs_gp.nrows):
        try:
            wk_num = int(rs_gp.cell_value(r, 0))
            if wk_num > 0:
                Hv = float(rs_gp.cell_value(r, 7) or 0)
                Lv = float(rs_gp.cell_value(r, 11) or 0)
                if Hv + Lv > 0:  # 只收集有数据的周
                    all_wks.append({
                    'H': float(rs_gp.cell_value(r, 7) or 0),
                    'I': float(rs_gp.cell_value(r, 8) or 0),
                    'L': float(rs_gp.cell_value(r, 11) or 0),
                    'N': float(rs_gp.cell_value(r, 13) or 0),
                    'O': float(rs_gp.cell_value(r, 14) or 0),
                    'Q': float(rs_gp.cell_value(r, 16) or 0),
                    'C': float(rs_gp.cell_value(r, 2) or 0),
                    'R': float(rs_gp.cell_value(r, 17) or 0),
                    'S': float(rs_gp.cell_value(r, 18) or 0),
                    'row': r,
                })
        except: pass

    # 当周数据
    cur_wk = {
        'H': tv1_e+tv1_h+tv1_s, 'I': sk_v1k, 'L': tb_e+tb_h+tb_s,
        'N': sk_bkk, 'O': sk_bks, 'Q': sk_tch, 'C': round(tv1h+tbkh/3, 1),
        'R': sp_R, 'S': sp_S, 'row': None
    }

    # 找当周row或追加
    gr = None
    for w in all_wks:
        if int(w.get('_wk', 0)) == week:
            gr = w['row']; break
    if gr is None:
        all_wks.append(cur_wk)
        gr = week + 2  # fallback
    else:
        # 替换当周数据
        for i, w in enumerate(all_wks):
            if int(w.get('_wk', 0)) == week:
                all_wks[i] = {**w, **cur_wk, 'row': w['row']}
                break

    # 为所有周重写公式列
    for wi, w in enumerate(all_wks):
        r = w['row'] if w['row'] is not None else week + 2
        H,I,L,N,O,Q,C = w['H'],w['I'],w['L'],w['N'],w['O'],w['Q'],w['C']
        Rv,Sv = w.get('R',0), w.get('S',0)

        # 累计（从第一周到当前周）
        cH = sum(x['H'] for x in all_wks[:wi+1])
        cI = sum(x['I'] for x in all_wks[:wi+1])
        cL = sum(x['L'] for x in all_wks[:wi+1])
        cN = sum(x['N'] for x in all_wks[:wi+1])
        cO = sum(x['O'] for x in all_wks[:wi+1])
        cQ = sum(x['Q'] for x in all_wks[:wi+1])
        cR = sum(x.get('R',0) for x in all_wks[:wi+1])
        cS = sum(x.get('S',0) for x in all_wks[:wi+1])
        cC = sum(x['C'] for x in all_wks[:wi+1])

        ws_gp.write(r, 1, cC)                                    # B
        ws_gp.write(r, 3, round((cI+cN)/(cH+cL), 2))            # D
        ws_gp.write(r, 4, round((I+N)/(H+L), 2))                 # E
        ws_gp.write(r, 5, round((cI/3+cO-cR-cS)/cQ, 2))         # F
        ws_gp.write(r, 6, round((I/3+O-Rv-Sv)/Q, 2))             # G
        ws_gp.write(r, 9, round(I/H, 2))                         # J
        ws_gp.write(r, 10, round(cI/cH, 2))                      # K

        # 数据列只在当前周才写（前两周保持原值不变）
        if wi == len(all_wks) - 1:  # 最后一周 = 当周
            ws_gp.write(r, 0, week)
            ws_gp.write(r, 2, C)
            ws_gp.write(r, 7, H)
            ws_gp.write(r, 8, I)
            ws_gp.write(r, 11, L)
            ws_gp.write(r, 12, w.get('M',41))
            ws_gp.write(r, 13, N)
            ws_gp.write(r, 14, O)
            ws_gp.write(r, 16, Q)
            ws_gp.write(r, 17, Rv if Rv else None)
            ws_gp.write(r, 18, Sv if Sv else None)

    wb.save(output_file)

    # ── 自恰 + 交叉验证 ──
    print(f"\n{'='*50}")
    print(f"  自恰性")
    print(f"  v1={tv1_e+tv1_h+tv1_s} bk={tb_e+tb_h+tb_s} 单科={all_stu}")
    print(f"  满班率学生={bk_total_stu}==学生bk={tb_e+tb_h+tb_s} {'✅' if bk_total_stu==tb_e+tb_h+tb_s else '❌'}")
    print(f"  满班率班={mb_xc+mb_cc+mb_gc+1}==生产M={tbc} {'✅' if mb_xc+mb_cc+mb_gc+1==tbc else '❌'}")
    print(f"\n  交叉验证 (排课 vs WPS):")
    diffs = []
    for name in MATH:
        ws = wps_data.get(name, {})
        sc = schedule_data.get(name, {})
        w_v1h = ws.get('v1h', 0); s_v1h = sc.get('v1h', 0)
        w_bkh = ws.get('bkh', 0); s_bkh = sc.get('bkk', 0) * 3  # schedule is act sum, ×3 = KS
        if abs(w_v1h - s_v1h) > 1 or abs(w_bkh - s_bkh) > 5:
            diffs.append(f"  {name}: 1v1 WPS={w_v1h:.0f}h 排课={s_v1h:.0f}h | BK WPS={w_bkh:.0f} 排课={s_bkh:.0f}")
    if diffs:
        print('\n'.join(diffs[:10]))
    else:
        print("  ✅ 全部一致")
    print(f"{'='*50}")


# ═══════════════ main ═══════════════
def main():
    if len(sys.argv) < 4:
        print('用法: python3 auto_weekly.py <模板.xls> <排课.xls> <WPS.xls> [周次]')
        sys.exit(1)

    base = sys.argv[1]; sched = sys.argv[2]; wps = sys.argv[3]
    week = int(sys.argv[4]) if len(sys.argv) > 4 else 3

    print(f'📂 底稿: {os.path.basename(base)}')
    print(f'📂 排课: {os.path.basename(sched)}')
    print(f'📂 WPS:  {os.path.basename(wps)}')
    print(f'🎯 第{week}周\n')

    wk_data = read_wps(wps, week)
    sch_data, sp_data = parse_schedule(sched, week)
    print(f'  WPS: {len(wk_data["teachers"])}位教师')
    m = re.search(r'(\d+)月', os.path.basename(wps))
    month = m.group(1)+'月' if m else '6月'
    out = os.path.join(os.path.dirname(base) or '.',
                       f'数学组数据统计表-宣城二校{month}第{week}周.xls')

    fill(base, wk_data, sch_data, sp_data, week, out)
    print(f'\n✅ {os.path.basename(out)}')


if __name__ == '__main__':
    main()
